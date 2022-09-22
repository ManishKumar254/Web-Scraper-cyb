import requests, webbrowser
import re
import csv
from bs4 import BeautifulSoup
import time
import ssl
import openpyxl
from selenium import webdriver
from selenium.webdriver.common import keys
from webdriver_manager.chrome import ChromeDriverManager
ssl._create_default_https_context = ssl._create_unverified_context

#replace url_data to rename the file
csv_file = open('url_data.csv', 'w')
csv_writer = csv.writer(csv_file)
csv_writer.writerow(['URL', 'Email' ,'Link'])

# print("Searching...")
url_list = []
unique_url_list = []
email_list = []
link_list=[]
unique_link_list=[]
form_urls=[]
e = 0
pre_list=[]
def get_email(urls_list):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
    e = 0
    for i in urls_list:
        flag = 1
        try:
            soup = "null"
            print("https://" + i)
            sourceCode = requests.get("https://" + i, headers=headers,timeout=25)
            soup = BeautifulSoup(sourceCode.text, 'lxml')
        except Exception as e:
            pass
        try:
            email = re.findall("([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)", soup.text, re.I)
            #Add negative keywords here
            agency=re.findall("agency",soup.text,re.I)
            software=re.findall("software",soup.text,re.I)
            #Here you can change the maximum number of occurance of negative keywords
            if len(agency)>=3 or len(software)>=5:
                email_list.append("skipped")
                flag=0
                print("skipped")
            else:
                email_list.append(email[0])
                print(email[0])
                flag = 0
        except Exception as e:
            flag = 1
            pass
        if flag == 1:
            try:
                print("https://" + i + "/contact/")
                sourceCode = requests.get("https://" + i + "/contact/", headers=headers,timeout=25)
                soup = BeautifulSoup(sourceCode.text, 'lxml')
            except:
                pass
            try:
                email = re.findall("([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)", soup.text, re.I)
                email_list.append(email[0])
                print(email[0])
                flag = 0
            except Exception as e:
                pass
        if flag == 1:
            try:
                print("https://" + i + "/about-us/")
                sourceCode = requests.get("https://" + i + "/about-us/", headers=headers,timeout=25)
                soup = BeautifulSoup(sourceCode.text, "lxml")
            except:
                pass
            try:
                email = re.findall("([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)", soup.text, re.I)
                email_list.append(email[0])
                print(email[0])
                flag = 0
            except Exception as e:
                pass
        if flag == 1:
            try:
                print("https://" + i + "/contact-us/")
                sourceCode = requests.get("https://" + i + "/contact-us/", headers=headers,timeout=25)
                soup = BeautifulSoup(sourceCode.text, "lxml")
            except:
                pass
            try:
                email = re.findall("([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)", soup.text, re.I)
                email_list.append(email[0])
                print(email[0])
                flag = 0
            except Exception as e:
                pass
        if flag == 1:
            try:
                print("https://" + i + "/support/")
                sourceCode = requests.get("https://" + i + "/support", headers=headers,timeout=25)
                soup = BeautifulSoup(sourceCode.text, "lxml")
            except:
                pass
            try:
                email = re.findall("([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)", soup.text, re.I)
                email_list.append(email[0])
                print(email[0])
                flag = 0
            except Exception as e:
                email_list.append("None")
                pass
        if flag == 1 :
            form_urls.append(i)






def remove_duplicates():
    i=0
    print("flag1")
    while i<len(url_list):
        print("loop " + str(i))
        if url_list[i] not in unique_url_list  :
            print(i)
            unique_url_list.append(url_list[i])
            unique_link_list.append(link_list[i])
        i=i+1

def display_elements():
    j = 0
    while j < len(email_list):
        if email_list[j] != "None" :
            a = unique_url_list[j] + " " + email_list[j]
            print(a)
        j += 1


def add_to_csv():
    j = 0
    while j < len(email_list):
        if email_list[j] != "None" and email_list[j] not in pre_list and email_list[j]!="skipped":
            a = unique_url_list[j] + " " + email_list[j]
            csv_writer.writerow([unique_url_list[j], email_list[j], unique_link_list[j]])
        j += 1
    #l=0
    #csv_writer.writerow(["Skipped","skipped","skipped"])
    #while l<len(email_list):
    #   if email_list[l]=="skipped":
    #       csv_writer.writerow([unique_url_list[l], email_list[l], unique_link_list[l]])
    #   l+=1
    
    #uncomment the code above to include skipped urls in url_data.csv
            
    csv_file.close()

def add_to_csv2():
    l=0
    #to Rename form excel sheet
    csv_file=open('form_data.csv','w')
    csv_writer2=csv.writer(csv_file)
    csv_writer2.writerow(['URL','Contact-us Form'])
    while l<len(form_urls):
        csv_writer2.writerow(([form_urls[l],'Present']))
        l=l+1
    csv_file.close()





def get_url_random(count):
    print("Searching...")
    url_list.clear()
    unique_url_list.clear()
    email_list.clear()
    link_list.clear()
    print("Searching from google 1...")
    n = 0
    while n < count:
        #you can increase the sleep time (seconds)
        time.sleep(15)
        #Here you can change the keywords our program is looking for 
        keyword_one="recruit"
        keyword_two="applicant"
        keyword_three="admin"
        google_search = requests.get(
            "https://www.google.com/search?rlz=1C2CHBF_enIN884IN930&source=hp&ei=6UwIYPmlGtXFrQGJ7rSwCw&q=site%3A.com+"+keyword_one+"+"+keyword_two+"+"+keyword_three+"+-buy+-%22try+for+free%22&oq=site%3A.com+"+keyword_one+"+"+keyword_two+"+"+keyword_three+"+-buy+-%22try+for+free%22"+"&start=" + str(
                n))
        soup = BeautifulSoup(google_search.text, 'lxml')
        x = soup.select(' .kCrYT a')
        if n != 0:
            print(str(n) + " urls searched!!")
        i = 0
        while i < 10:
            try:
                y = x[i]['href']
                z = y.split('=')[1]
                z2 = z.split("/", 3)[2]
                url_list.append(z2)
                print(z)
                link_list.append(z)
                print(z2)
                print(str(len(url_list)) + " added!!")
            except:
                pass
            i += 1
        n += 10
    #print("Searching from google 2...")
    n = 0
    while n < count:
        #you can increase the sleep time (seconds)
        time.sleep(10)
        #Here you can change the keywords our program is looking for
        keyword_one="tracking"
        keyword_two="upgrade"
        keyword_three="system"
        google_search = requests.get(
            "https://www.google.com/search?rlz=1C2CHBF_enIN884IN930&source=hp&ei=6UwIYPmlGtXFrQGJ7rSwCw&q=site%3A.com+"+keyword_one+"+"+keyword_two+"+"+keyword_three+"+-buy+-%22try+for+free%22&oq=site%3A.com+"+keyword_one+"+"+keyword_two+"+"+keyword_three+"+-buy+-%22try+for+free%22"+"&start=" + str(
                n))
        soup = BeautifulSoup(google_search.text, 'lxml')
        x = soup.select(' .kCrYT a')
        if n != 0:
            print(str(n) + " urls searched!!")
        i = 0
        while i < 10:
            try:
                y = x[i]['href']
                z = y.split('=')[1]
                z2 = z.split("/", 3)[2]
                url_list.append(z2)
                print(z)
                link_list.append(z)
                print(z2)
                print(str(len(url_list)) + " added!!")
            except:
                pass
            i += 1
        n += 10
    print("Searching from google 3...")
    n = 0
    while n < count:
        #you can increase the sleep time (seconds)
        time.sleep(15)
        #Here you can change the keywords our program is looking for
        keyword_one="admin"
        keyword_two="upgrade"
        keyword_three="credify"
        google_search = requests.get(
            "https://www.google.com/search?rlz=1C2CHBF_enIN884IN930&source=hp&ei=6UwIYPmlGtXFrQGJ7rSwCw&q=site%3A.com+"+keyword_one+"+"+keyword_two+"+"+keyword_three+"+-buy+-%22try+for+free%22&oq=site%3A.com+"+keyword_one+"+"+keyword_two+"+"+keyword_three+"+-buy+-%22try+for+free%22"+"&start=" + str(
                n))
        soup = BeautifulSoup(google_search.text, 'lxml')
        x = soup.select(' .kCrYT a')
        if n != 0:
            print(str(n) + " urls searched!!")
        i = 0
        while i < 10:
            try:
                y = x[i]['href']
                z = y.split('=')[1]
                z2 = z.split("/", 3)[2]
                url_list.append(z2)
                print(z)
                link_list.append(z)
                print(z2)
                print(str(len(url_list)) + " added!!")
            except:
                pass
            i += 1
        n += 10


def get_url_random2(count):
    print("Searching...")
    url_list.clear()
    unique_url_list.clear()
    email_list.clear()
    link_list.clear()
    print("Searching from google 1...")
    n = 0
    while n < count:
        #you can increase the sleep time (seconds)
        time.sleep(15)
        #Here you can change the keywords our program is looking for
        keyword_one="business blogs"
        keyword_two="hr blogs"
        keyword_three="hiring"
        google_search = requests.get(
            "https://www.google.com/search?rlz=1C2CHBF_enIN884IN930&source=hp&ei=6UwIYPmlGtXFrQGJ7rSwCw&q=site%3A.com+"+keyword_one+"+"+keyword_two+"+"+keyword_three+"+-buy+-%22try+for+free%22&oq=site%3A.com+"+keyword_one+"+"+keyword_two+"+"+keyword_three+"+-buy+-%22try+for+free%22"+"&start=" + str(
                n))
        soup = BeautifulSoup(google_search.text, 'lxml')
        x = soup.select(' .kCrYT a')
        if n != 0:
            print(str(n) + " urls searched!!")
        i = 0
        while i < 10:
            try:
                y = x[i]['href']
                z = y.split('=')[1]
                z2 = z.split("/", 3)[2]
                url_list.append(z2)
                print(z)
                link_list.append(z)
                print(z2)
                print(str(len(url_list)) + " added!!")
            except:
                pass
            i += 1
        n += 10
    print("Searching from google 2...")
    n = 0
    while n < count:
        #you can increase the sleep time (seconds)
        time.sleep(10)
        #Here you can change the keywords our program is looking for
        keyword_one="blogs"
        keyword_two="talent acquisition"
        keyword_three="recruiting"
        google_search = requests.get(
            "https://www.google.com/search?rlz=1C2CHBF_enIN884IN930&source=hp&ei=6UwIYPmlGtXFrQGJ7rSwCw&q=site%3A.com+"+keyword_one+"+"+keyword_two+"+"+keyword_three+"+-buy+-%22try+for+free%22&oq=site%3A.com+"+keyword_one+"+"+keyword_two+"+"+keyword_three+"+-buy+-%22try+for+free%22"+"&start=" + str(
                n))
        soup = BeautifulSoup(google_search.text, 'lxml')
        x = soup.select(' .kCrYT a')
        if n != 0:
            print(str(n) + " urls searched!!")
        i = 0
        while i < 10:
            try:
                y = x[i]['href']
                z = y.split('=')[1]
                z2 = z.split("/", 3)[2]
                url_list.append(z2)
                print(z)
                link_list.append(z)
                print(z2)
                print(str(len(url_list)) + " added!!")
            except:
                pass
            i += 1
        n += 10
    print("Searching from google 3...")
    n = 0
    while n < count:
        #you can increase the sleep time (seconds)
        time.sleep(15)
        #Here you can change the keywords our program is looking for
        keyword_one="blogs"
        keyword_two="applicant tracking system"
        keyword_three="executive search"
        google_search = requests.get(
            "https://www.google.com/search?rlz=1C2CHBF_enIN884IN930&source=hp&ei=6UwIYPmlGtXFrQGJ7rSwCw&q=site%3A.com+"+keyword_one+"+"+keyword_two+"+"+keyword_three+"+-buy+-%22try+for+free%22&oq=site%3A.com+"+keyword_one+"+"+keyword_two+"+"+keyword_three+"+-buy+-%22try+for+free%22"+"&start=" + str(
                n))
        soup = BeautifulSoup(google_search.text, 'lxml')
        x = soup.select(' .kCrYT a')
        if n != 0:
            print(str(n) + " urls searched!!")
        i = 0
        while i < 10:
            try:
                y = x[i]['href']
                z = y.split('=')[1]
                z2 = z.split("/", 3)[2]
                url_list.append(z2)
                print(z)
                link_list.append(z)
                print(z2)
                print(str(len(url_list)) + " added!!")
            except:
                pass
            i += 1
        n += 10
def previous_list():
    loc="previous_list.xlsx"
    wb=openpyxl.load_workbook(loc)
    sheet_obj=wb.active
    url_list=[]
    j=1
    while j<sheet_obj.max_row+2:
        cell_obj=sheet_obj.cell(row=j,column=2)
        url_list.append(cell_obj.value)
        j=j+2
    for i in url_list:
        pre_list.append(i)
        print(i)
    print(pre_list)


print(previous_list())
y = int(input("Enter the no. of urls to be scanned: "))
print(get_url_random(y))
print(get_url_random2(y))
print("Starting")
print(link_list)
remove_duplicates()
print("done")
get_email(unique_url_list)
add_to_csv()
add_to_csv2()
display_elements()
print()
print("Urls without any email address mentioned on the webite have been removed!")
print(str(len(email_list)) + " urls Found!!")
print("Note: All the data has been logged into a csv file 'data.csv'.")
print()


